# AEFBookReport.py

# Load the openpyxl Excel library
from openpyxl import load_workbook
from openpyxl.styles import Font


class AEFBookReport:

	def __init__(self, str_title):
		self.str_title      = str_title
		self.sheet_reports  = []
		self.str_success    = "Syntax check completed successfully."
		self.str_fail       = "Syntax check found errors."
		self.is_valid       = True

	def add_sheet_report(self, sheet_report):
		self.sheet_reports.append(sheet_report)
		return True


	# def is_valid(self, is_valid):
	# 	self.is_valid   = is_valid


	def print(self, workbook):
		results_sheet	= workbook.create_sheet(title = "Syntax check results", index = 7)
		results_sheet.cell(1, 1, value=self.str_title)
		results_sheet.cell(1, 1).font	= Font(bold=True)
	
		x_row	= 2
		for sheet_report in self.sheet_reports:
			x_row	= sheet_report.print(workbook, results_sheet, x_row)
		x_row	+= 1

		if (self.is_valid):
			str_result  = self.str_success
		else:
			str_result  = self.str_fail
		results_sheet.cell(x_row, 1, value=str_result)

	