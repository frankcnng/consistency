# aefsheets.py

from openpyxl.comments import Comment
import re

from syntaxreport.AEFSheetReport import AEFSheetReport
from syntaxreport.AEFCellReport import AEFCellReport


class AEFSheet:
	"""Abstract superclass that checks the syntax of a worksheet."""

	def __init__(self, worksheet, field_names):
		self.worksheet		= worksheet
		self.field_names	= field_names


	def check_structure(self, check_report):
		return self.check_field_names(check_report)


	def check_fields_dimensions(self, x_heading, x_start, x_end, n_blanks, worksheet, template_fields, sheet_report):
		n_template_fields	= len(template_fields)
		if (x_heading == 0):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Could not find '" + template_fields[0] + "' section in worksheet.")
			return False
		elif (x_start == 0):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Could not find '" + template_fields[1] + "' field in worksheet.")
			return False
		elif (x_end == 0):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Could not find '" + template_fields[n_template_fields - 1] + "' field in worksheet.")
			return False
		elif ((x_end - x_start - n_blanks + 1) != n_template_fields - 1):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Number of fields in '" + template_fields[0] + "' worksheet is incorrect.")
			return False
		return True


	# def check_field_names(self, sheet_report):
	# 	return True


	# def check_content(self, sheet_report):
	# 	return True
	

	def check_cell_content(self, x_target_row, x_target_column, x_tuple, sheet_report):
		field_reg_exp_tuple	= self.field_reg_exp_tuples[x_tuple]
		field_name			= field_reg_exp_tuple[0]
		field_reg_exp		= field_reg_exp_tuple[1]
		if (field_reg_exp == ""):
			return True
		field_error_mesg	= field_reg_exp_tuple[2]
		cell				= self.worksheet.cell(x_target_row, x_target_column)
		# if the cell can be either blank, of a defined set of values/regexp
		if ((re.match("^blankable", field_reg_exp) != None) and (cell.value == None)):
			return True

		if (cell.data_type == 'd'):
#			if (re.match(field_reg_exp, str(cell.number_format)) == None):
			if (cell.is_date is False):
				str_message	= "Cell content error: The value provided for '" + field_name + field_error_mesg
				sheet_report.add_cell_report(self.template_sheet_name, cell, str_message)
				return False
		elif (re.fullmatch(field_reg_exp, str(cell.value))) == None:
			str_message	= "Cell content error: The value provided for '" + field_name + field_error_mesg
			sheet_report.add_cell_report(self.template_sheet_name, cell, str_message)
			return False
		return True

	def add_comment(self, cell, str_comment):
		comment	= Comment(str_comment, "AEFSyntaxCheck")
		cell.comment	= comment


	def coord2cell_ref(self, x_row, x_column):
		# return the alpha number excel cell reference from row number and column number (1-based)
		if (x_column <= 26):
			return (chr(ord('A') + x_column - 1) + str(x_row))
		else:
			return ("A" + chr(ord('A') + (x_column % 26) - 1) + str(x_row))


class RowFieldsSheet(AEFSheet):
	"""Abstract superclass for checks syntax of worksheets with fields organised in rows."""

	def check_field_names(self, book_report):
	# Check the names of the field names (headings) in the sheet correspond with those in the field_names array

		heading_column		= 0
		fields_start_column	= 0
		fields_end_column	= 0
		field_headings_row	= 0
		template_fields		= self.field_names
		n_template_fields	= len(template_fields)
		worksheet			= self.worksheet
		sheet_report	= AEFSheetReport(self.template_sheet_name, 3)
		book_report.add_sheet_report(sheet_report)
		for column in worksheet.iter_rows(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
			if (field_headings_row > 0):
				break
			else:
				n_blank_cells	= 0
			for cell in column:
				if ((cell.data_type == "s") and (cell.value.casefold() == template_fields[0].casefold())):
					heading_column	= cell.column
					continue
				elif ((heading_column > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[1].casefold())):
					fields_start_column	= cell.column
					field_headings_row	= cell.row
					continue
				elif (fields_start_column > 0):
					if (cell.row == field_headings_row):
						if ((cell.value == None)):
							n_blank_cells	= n_blank_cells + 1
							continue
						elif ((fields_start_column > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[n_template_fields - 1].casefold())):
							fields_end_column	= cell.column
							break
					else:
						break

		if (self.check_fields_dimensions(heading_column, fields_start_column, fields_end_column, n_blank_cells, worksheet, template_fields, sheet_report)) is False:
			return False

		dest_fields	= []
		for column in worksheet.iter_rows(min_col=fields_start_column, max_col=fields_end_column, min_row=field_headings_row, max_row=field_headings_row, values_only=True):
			dest_fields.append(column)

		is_valid	= True
		for x_field in range (1, n_template_fields):
			if (template_fields[x_field] in dest_fields[0]) is False:
				sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "The field '" + template_fields[x_field] + "' cannot be found in worksheet")
				is_valid	= False
		if (is_valid):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "All fields found.")
		return is_valid


	def check_content(self, check_report):
	# Check the contents of the fields in the sheet are syntactically correct
	
		sheet_report	= AEFSheetReport(self.template_sheet_name, 3)
		check_report.add_sheet_report(sheet_report)

		fields_start_column	= 0
		fields_end_column	= 0
		fields_row			= 0
		template_fields		= self.field_names
		n_template_fields	= len(template_fields)
		worksheet			= self.worksheet
		for column in worksheet.iter_rows(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
			for cell in column:
				if ((cell.data_type == "s") and (cell.value.casefold() == template_fields[1].casefold())):
					fields_start_column	= cell.column
					fields_row			= cell.row
					continue
				elif ((fields_start_column > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[n_template_fields - 1].casefold())):
					fields_end_column	= cell.column
					break

		x_row		= fields_row + 1	# content is in the row after the field names
		is_valid	= True
		is_empty	= False
		while (is_empty) is False:
			is_empty	= True
			for x_column in range(fields_start_column, fields_end_column):
				if (worksheet.cell(x_row, x_column).value != None):
					is_empty	= False
			x_row	+= 1
		x_last_row	= x_row - 1


		is_valid	= True
		for x_row in range(fields_row + 1, x_last_row):
			is_empty	= True
			x_tuple		= 0
			for x_column in range(fields_start_column, fields_end_column):
				# if (worksheet.cell(x_row, x_column).value != None):
				# 	is_empty	= False
				if (self.check_cell_content(x_row, x_column, x_tuple, sheet_report)) is False:
					is_valid	= False
				x_tuple	+= 1
			x_row	+= 1

		if (is_valid):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "All field content valid.")
		return is_valid


class AEFAuthorizations(RowFieldsSheet):
	"""Concrete subclass for AEF Authorizations worksheet."""

	def __init__(self, worksheet, field_names):
		self.template_sheet_name	= "Table 2 Authorizations"

		self.field_reg_exp_tuples	=	[
											["Authorization ID", "[A-Za-z0-9 \-]+", "' can only contain alphanumeric, space, and hyphen characters."],
											["Date of authorization", "dd/mm/yyyy", "'' must be a valid date."],
											["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
											["Version of the authorization", "\d+", "' must be a number."],
											["", "", ""],
											["Authorized quantity", "\d+", "' must be a number."],
											["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
											["Applicable GWP value(s)", "", ""],
											["Applicable non-GHG metric", "", ""],
											["Sector(s)", "[A-Za-z0-9 ]+", "' can only contain alphanumeric, and space characters."],
											["Activity type(s)", "[A-Za-z0-9 \+]+", "' can only contain alphanumeric, space, and '+' characters."],
											["Purposes for authorization", "NDC|OIMP|IMP|OP|NDC and OIMP|NDC and IMP|NDC and OP", "' must be one of 'NDC', 'OIMP', 'IMP', 'OP', 'NDC and OIMP', 'NDC and IMP', or 'NDC and OP'."],
											["Authorized Party(ies) ID", "[A-Z]{3}( *, *[A-Z]{3})*", "' must a comma-separated list of ISO 3166 alpha-3 codes."],
											["Authorized entity(ies) ID", "[A-Za-z0-9 \-\.\(\)]+( *, *[A-Za-z0-9 \-\.\(\)]+)*", "' must be a comma-separated list of entity names."],
											["OIMP authorized by the Party", "", ""],
											["Authorized timeframe", "blankable|^ *(?:Occurred|Use): from\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))\s+to\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))(?:(?:\.\s*Use:\s+from\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))\s+to\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4})))?) *$",\
												"' must be empty or 'Occurred: from <date> to <date>', 'Use: from <date> to <date>', or 'Occurred: from <date> to <date>. Use: from <date> to <date>'\n where <date> is 'yyyy', 'mm/yyyy', or 'dd/mm/yyyy'."],
											["Authorization terms and conditions", "", ""],
											["Authorization documentation", "", ""],
											["First transfer definition for OIMP", "blankable|NA|Authorization|Issuance|Use of cancellation", "' must be empty or one of 'NA', 'Authorization', 'Issuance', 'Use of cancellation;."],
											["Additional explanatory information", "", ""]
										]

		super().__init__(worksheet, field_names)


class AEFActions(RowFieldsSheet):
	"""Concrete subclass for AEF Actions worksheet."""

	def __init__(self, worksheet, field_names):
		self.template_sheet_name	= "Table 3 Actions"

		self.field_reg_exp_tuples	=	[
											["Action date", "dd/mm/yyyy", "'' must be a valid date."],
											["Action type", "Acquisition|Transfer|Use|Cancellation|First transfer", "'' must be one of 'Acquisition', 'Transfer', 'Use', 'Cancellation', 'First transfer'"],
											["Action subtype", "", ""],
											["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
											["Authorization ID", "[A-Za-z0-9 \-]+", "' can only contain alphanumeric, space, and hyphen characters."],
											["First transferring participating Party ID", "[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
											["Party ITMO registry ID", "[A-Z]{3}\d{2}", "' must be a Party ID followed by two digits"],
											["First ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
											["Last ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
											["", "", ""],
											["Underlying unit registry ID", "", ""],
											["First unit ID", "", ""],
											["Last unit ID", "", ""],
											["", "", ""],
											["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
											["Applicable GWP value(s)", "", ""],
											["Applicable non-GHG metric", "", ""],
											["Quantity (t CO2 eq)", "blankable|\d+", "' must be a numerical value."],
											["Quantity (in non-GHG metric)", "blankable|\d*", "' must be a numerical value."],
											["", "", ""],
											["Mitigation type", "", ""],
											["Vintage", "blankable|\d{4}", "' must be a year."],
											["", "", ""],
											["Transferring participating Party ID", "blankable|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
											["Acquiring participating Party ID",  "blankable|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
											["", "", ""],
											["Purpose for which the ITMO has been used towards or cancelled for OIMP", "", ""],
											["Using/cancelling participating Party ID", "blankable|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
											["Using/cancelling authorized entity ID", "", ""],
											["Calendar year for which the ITMOs are used towards the Party's NDC", "blankable|\d{4}", "' must be a year."],
											["", "", ""],
											["Result of the consistency checks", "", ""],
											["Additional explanatory information", "", ""]
										]

		super().__init__(worksheet, field_names)


class AEFHoldings(RowFieldsSheet):
	"""Concrete subclass for AEF Holdings worksheet."""

	def __init__(self, worksheet, field_names):
		self.template_sheet_name	= "Table 4 Holdings"

		self.field_reg_exp_tuples	=	[
											["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
											["Authorization ID", "[A-Za-z0-9 \-]+", "' can only contain alphanumeric, space, and hyphen characters."],
											["First transferring participating Party ID", "[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
											["Party ITMO registry ID", "[A-Z]{3}\d{2}", "' must be a Party ID followed by two digits"],
											["First ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
											["Last ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
											["", "", ""],
											["Underlying unit registry ID", "", ""],
											["First unit ID", "", ""],
											["Last unit ID", "", ""],
											["", "", ""],
											["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
											["Applicable GWP value(s)", "", ""],
											["Applicable non-GHG metric", "", ""],
											["Quantity (t CO2 eq)", "blankable|NA|\d+", "' must be a numerical value."],
											["Quantity (in non-GHG metric)", "blankable|NA|\d+", "' must be a numerical value."],
											["", "", ""],
											["Mitigation type", "", ""],
											["Vintage", "d{4}", "' must be a year."]
										]

		super().__init__(worksheet, field_names)


class AEFAuthEntities(RowFieldsSheet):
	"""Concrete subclass for AEF Authorized Entities worksheet."""

	def __init__(self, worksheet, field_names):
		self.template_sheet_name	= "Table 5 Auth. entities"

		self.field_reg_exp_tuples	=	[
											["Date of the authorization", "blankable|dd/mm/yyyy", "'' must be a valid date."],
											["Name", "", ""],
											["Country of incorporation", "blankable|[A-Z][A-Za-z \(\)\']+", "'' is not a recognised Party Name."],	# Capitalised alphabet string contains spaces, brackets, apostrophes],
											["Identification number", "", ""],
											["Cooperative approach ID", "blankable|CA\d{4}", "' must start with 'CA' followed by four digits."],
											["Conditions", ""],
											["Change and revocation conditions", "", ""],
											["Additional explanatory information", "", ""]
										]

		super().__init__(worksheet, field_names)


class ColumnFieldsSheet(AEFSheet):
	"""Abstract superclass for checking worksheets with fields organised in columns."""


	def check_field_names(self, book_report):
		heading_row			= 0
		fields_start_row	= 0
		fields_end_row		= 0
		fields_column		= 0
		template_fields		= self.field_names
		n_template_fields	= len(template_fields)
		worksheet			= self.worksheet
		str_subheading		= self.template_sheet_name
		if (self.template_sheet_name == "Summary information"):
			str_subheading	+= ": " + template_fields[0]
		sheet_report		= AEFSheetReport(str_subheading, 3)
		book_report.add_sheet_report(sheet_report)
		for row in worksheet.iter_cols(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
			for cell in row:
				if ((cell.data_type == "s") and (cell.value.casefold() == template_fields[0].casefold())):
					heading_row	= cell.row
					fields_column	= cell.column
					continue
				elif ((heading_row > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[1].casefold())):
					fields_start_row	= cell.row
					continue
				elif ((fields_start_row > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[n_template_fields - 1].casefold())):
					fields_end_row	= cell.row
					break
		if (self.check_fields_dimensions(heading_row, fields_start_row, fields_end_row, 0, worksheet, template_fields, sheet_report)) is False:
			return False

		dest_fields	= []
		for row in worksheet.iter_cols(min_col=fields_column, max_col=fields_column, min_row=fields_start_row, max_row=fields_end_row, values_only=True):
			dest_fields.append(row)

		for x_field in range (1, n_template_fields):
			if (template_fields[x_field] in dest_fields[0]) is False:
				sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Could not find '" + template_fields[x_field] + "' field in worksheet.")
				return False
		sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "All fields found.")
		return True


class AEFSummary(ColumnFieldsSheet):

	def __init__(self):
		self.template_sheet_name	= "Summary information"
		self.submission_fields		=	[
											"Table 1: Submission",
											"Party",
											"Version",
											"Reported year",
											"Date of submission",
											"Review status of the initial report",
											"Result of the consistency check of this AEF submission",
											"First year of the NDC implementation period",
											"Last year of the NDC implementation period",
											"Reference to the Article 6 technical expert review report of the initial report"
										]
		self.authorizations_fields =	[
											"Table 2: Authorizations",
											"Authorization ID",
											"Date of authorization",
											"Cooperative approach ID",
											"Version of the authorization",
											"Authorized quantity",
											"Metric",
											"Applicable GWP value(s)",
											"Applicable non-GHG metric",
											"Sector(s)",
											"Activity type(s)",
											"Purposes for authorization",
											"Authorized Party(ies) ID",
											"Authorized entity(ies) ID",
											"OIMP authorized by the Party",
											"Authorized timeframe",
											"Authorization terms and conditions",
											"Authorization documentation",
											"First transfer definition for OIMP",
											"Additional explanatory information"
										]
		self.actions_fields			=	[
											"Table 3: Actions",
											"Action date",
											"Action type",
											"Action subtype",
											"Cooperative approach ID",
											"Authorization ID",
											"First transferring participating Party ID",
											"Party ITMO registry ID",
											"First ID",
											"Last ID",
											"Underlying unit registry ID",
											"First unit ID",
											"Last unit ID",
											"Metric",
											"Applicable GWP value(s)",
											"Applicable non-GHG metric",
											"Quantity (t CO2 eq)",
											"Quantity (in non-GHG metric)",
											"Mitigation type",
											"Vintage",
											"Transferring participating Party ID",
											"Acquiring participating Party ID",
											"Purpose for which the ITMO has been used towards or cancelled for OIMP",
											"Using/cancelling participating Party ID",
											"Using/cancelling authorized entity ID",
											"Calendar year for which the ITMOs are used towards the Party's NDC",
											"Result of the consistency checks",
											"Additional explanatory information"
										]
		self.holdings_fields		=	[
											"Table 4: Holdings",
											"Cooperative approach ID",
											"Authorization ID",
											"First transferring participating Party ID",
											"Party ITMO registry ID",
											"First ID",
											"Last ID",
											"Underlying unit registry ID",
											"First unit ID",
											"Last unit ID",
											"Metric",
											"Applicable GWP value(s)",
											"Applicable non-GHG metric",
											"Quantity (t CO2 eq)",
											"Quantity (in non-GHG metric)",
											"Mitigation type",
											"Vintage"
										]
		self.auth_entities_fields	=	[
											"Table 5: Authorized entities",
											"Date of the authorization",
											"Name",
											"Country of incorporation",
											"Identification number",
											"Cooperative approach ID",
											"Conditions",
											"Change and revocation conditions",
											"Additional explanatory information"
										]


	def check_structure(self, workbook, field_names, check_report):
		self.set_field_names(field_names)
		self.worksheet	= workbook[self.template_sheet_name]
		for x_table in range (0, 4):
			self.field_names	= field_names[x_table]
			if (self.check_field_names(check_report)) is False:
				return False
		return True


	def set_field_names(self, field_names):
		field_names.append(self.submission_fields)
		field_names.append(self.authorizations_fields)
		field_names.append(self.actions_fields)
		field_names.append(self.holdings_fields)
		field_names.append(self.auth_entities_fields)


class AEFSubmission(ColumnFieldsSheet):
	def __init__(self, worksheet, field_names):
		self.template_sheet_name	= "Table 1 Submission"
		self.field_reg_exp_tuples	=	[
											["Party", "[A-Z][A-Za-z \(\)\']+", "'' is not a recognised Party Name."],	# Capitalised alphabet string contains spaces, brackets, apostrophes
											["Version", "[0-9]+\.[0-9]+", "' must conform to X.Y."],
											["Reported year", "\d{4}", "'' must be a four digit year."],
											["Date of submission", "dd/mm/yyyy", "'' must be a valid date"],
											["Review status of the initial report", "\{Information in this field is populated by the CARP\}", "' must not be changed.  It is for secretariat use."],
											["Result of the consistency check of this AEF submission", "\{Information in this field is populated by the CARP\}", "'' must not be changed.  It is for secretariat use."],
											["First year of the NDC implementation period", "\d{4}", "'' must be a four digit year."],
											["Last year of the NDC implementation period", "\d{4}", "'' must be a four digit year."],
											["Reference to the Article 6 technical expert review report of the initial report", "\{Link to be produced by the CARP\}", "'' must not be changed.  It is for secretariat use."]
										]
		super().__init__(worksheet, field_names)


	def check_content(self, check_report):
		sheet_report	= AEFSheetReport(self.template_sheet_name, 3)
		check_report.add_sheet_report(sheet_report)

		fields_start_row	= 0
		fields_end_row		= 0
		fields_column		= 0
		template_fields		= self.field_names
		n_template_fields	= len(template_fields)
		worksheet			= self.worksheet
		for row in worksheet.iter_cols(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
			for cell in row:
				if ((cell.data_type == "s") and (cell.value.casefold() == template_fields[1].casefold())):
					fields_start_row	= cell.row
					fields_column		= cell.column
					continue
				elif ((fields_start_row > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[n_template_fields - 1].casefold())):
					fields_end_row	= cell.row
					break

		column		= fields_column + 1	# content is in the column after the field names
		x_tuple		= 0
		is_valid	= True
		for x_row in range(fields_start_row, fields_end_row):
			if (self.check_cell_content(x_row, column, x_tuple, sheet_report)) is False:
				is_valid	= False
			x_tuple	+= 1
		if (is_valid):
			sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "All field content valid.")
		return is_valid
