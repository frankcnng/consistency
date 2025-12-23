# main.py
#
# Checks the syntax of AEF files
#

# Load os and shell libraries
import os, shutil
import time

# Load the openpyxl Excel library
from openpyxl import load_workbook

from aef_structure_check import AEFStructureCheck
from aef_content_check import AEFContentCheck
from syntaxreport.AEFBookReport import AEFBookReport


# The directory containing AEF files to be checked
aef_dir			= "/Users/frankng/United Nations Framework Convention on Climate Change/Mitigation - Article 6.2/03_CARP/consistency/AEF_files/"
unprocessed_dir	= aef_dir + "00.unprocessed/"
archive_dir		= aef_dir + "99.archive/"
passed_dir		= aef_dir + "10.syntax.passed/"
failed_dir		= aef_dir + "11.syntax.failed/"
#aef_dir		= "../AEF_files.00.unchecked/"


def main():
	"""	Check all .xlsx files in unprocessed_dir for valid AEF syntax.
		Files ending with '.syntax_passed.xlsx' or '.syntax_failed.xlsx' are the output of this tool, and will be overwritten.
	"""
	files	= os.listdir(unprocessed_dir)
	for str_file in files:
		if (str_file.endswith(".xlsx")):
			if str_file.endswith(".syntax_checked.xlsx"):
				continue
			else:
				str_head	= str_file[:-5]
				str_checked	= str_head + ".syntax_checked.xlsx"
				str_failed	= str_head + ".syntax_failed.xlsx"
				dst_path	= unprocessed_dir + str_checked
				src_path	= unprocessed_dir + str_file
				shutil.copyfile(src_path, dst_path)
				str_submission_key, is_valid	= check_file(dst_path, str_file)
				if (is_valid):
					src_path	= dst_path
					dst_path	= passed_dir + str_submission_key + ".syntax_passed.xlsx"
					if (os.path.exists(dst_path)):
						dst_path	= failed_dir + str_failed
						shutil.move(src_path, dst_path)							# move source to destination
						str_error	= "Duplicate submission.  Submission with the same Party ID (" + str_submission_key[0:3] + "), Reported Year (" + str_submission_key[4:8] + "), Version (" + str_submission_key[9:10] + "." + str_submission_key[11:12] + ") already submitted."
						update_submission_status(dst_path, str_error)
					else:
						shutil.move(src_path, dst_path + ".syntax_passed.xlsx")
				else:
					shutil.move(dst_path, failed_dir + str_failed)
				shutil.move(unprocessed_dir + str_file, archive_dir + str_file)


def update_submission_status(path, str_error):
	"""	Update the worksheet status for the submission file at 'path',
		move the file to directory 'str_dir', and set the error message to 'str_error'.
	"""
	workbook	= load_workbook(path, data_only=True)
	worksheet	= workbook["Table 1 Submission"]
	worksheet.cell(9, 3, value=str_error + ". " + time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()) + " GMT")
	workbook.save(path)


def check_file(str_path, str_file):
	"""	Check the file with pathname 'str_path'.
		The local name of the file is 'str_file'
		First the structure of the workbook is checked,
		next, the content of the fields in each sheet are checked.
	"""
	workbook			= load_workbook(str_path)
	worksheets			= []
	field_names			= []
	check_report		= AEFBookReport(str_file)
	str_submission_key	= ""

	structureCheck = AEFStructureCheck()
	if structureCheck.check(workbook, worksheets, field_names, check_report) is False:
		check_report.is_valid = False
	else:
		contentCheck 								= AEFContentCheck()
		str_submission_key, check_report.is_valid	= contentCheck.check(worksheets, field_names, check_report)
	check_report.print(workbook)
	
	workbook.save(str_path)
	return str_submission_key, check_report.is_valid


if __name__ == "__main__":
	main()
