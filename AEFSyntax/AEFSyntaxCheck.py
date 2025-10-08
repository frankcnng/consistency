# AEFSyntaxCheck.py
#
# Checks the syntax of AEF files
#

# Load os and shell libraries
import os, shutil

# Load the openpyxl Excel library
from openpyxl import load_workbook

from AEFStructureCheck import AEFStructureCheck
from AEFContentCheck import AEFContentCheck


# The directory containing AEF files to be checked
#aef_dir	= "/Users/frankng/United Nations Framework Convention on Climate Change/Mitigation - Article 6.2/03_CARP/consistency/AEF_files/"
aef_dir	= "../AEF_files/"


def main():
	files	= os.listdir(aef_dir)
	for file in files:
		if file.endswith(".syntax_checked.xlsx"):
			continue
		else:
			head		= file[:-5]
			dst_path	= aef_dir + head + ".syntax_checked.xlsx"
			src_path	= aef_dir + file
			shutil.copyfile(src_path, dst_path)

			print ("\nChecking '" + file + "'")
			check_file(dst_path)


def check_file(file):
	workbook 		= load_workbook(aef_dir + file)
	results_sheet	= workbook.create_sheet(title = "Syntax check results", index = 7)
	worksheets		= []
	field_names		= []
	str_results		= [""]

	structureCheck = AEFStructureCheck()
	if structureCheck.check(workbook, worksheets, field_names, str_results) is False:
		str_results[0]	+= "\nSyntax check found errors"
	else:
		contentCheck = AEFContentCheck()
		if contentCheck.check(worksheets, field_names, str_results) is False:
			str_results[0]	+= "\nSyntax check found errors."
		else:
			str_results[0]	+= "\nSyntax check completed successfully."

	print (str_results[0])
	results_sheet.cell(1,1, value=str_results[0])
	
	workbook.save(aef_dir + file)


if __name__ == "__main__":
	main()
