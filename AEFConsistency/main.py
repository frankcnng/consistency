# main.py
#
# Checks the consistency of AEF files
#

# Load os and shell libraries
import os, shutil

# Load the openpyxl Excel library
from openpyxl import load_workbook
import sqlite3

import aef_sheet
import aef_submission
from aef_consistency_check.AEFConsistencyReport import *
from create_db import create_tables


# The directory containing AEF files to be checked
aef_dir				= "/Users/frankng/United Nations Framework Convention on Climate Change/Mitigation - Article 6.2/03_CARP/consistency/AEF_files/"
syntax_passed_dir	= aef_dir + "10.syntax.passed/"
consistent_dir		= aef_dir + "20.consistent/"
inconsistent_dir	= aef_dir + "21.inconsistent/"
obsolete_dir		= aef_dir + "30.obsolete/"
duplicate_dir		= aef_dir + "31.duplicate/"

def main():
    """ Check all .xlsx files in aef_dir/syntax_passed_dir for AEF consistency.
        Files ending with '.consistency_checked.xlsx' are the output of this tool, and if any exist, will be overwritten.
    """
    conn    = create_tables(":memory:") #create an in-memory database for consistency checking
    cursor  = conn.cursor()
    load_submissions(consistent_dir, cursor)
    load_submissions(inconsistent_dir, cursor)
    load_submissions(syntax_passed_dir, cursor)

    replace_obsolete_submissions(cursor)
    unload_inconsistent_submissions(cursor)

    check_new_submissions(cursor)
    conn.close()
    return


def	load_submissions(str_path, cursor):
    """ Load AEF submission files in pathname 'str_path' into the database.
        The files have been syntax checked, so fields can be loaded into objects without checking.
    """
    files	= os.listdir(str_path)
    for str_file in files:
        if (str_file.endswith(".xlsx")):
            str_filepath    = str_path + str_file
            workbook	    = load_workbook(str_filepath, data_only=True)	# load workbook openpyxl object
            load_workbook_to_db(workbook, cursor, str_filepath)			    # load workbook into database
    return


def	load_workbook_to_db(workbook, cursor, str_filepath):
    """ Write all data sheets of workbook into the database.
        The workbook has been syntax checked, so fields can be loaded into objects without checking.
    """
    try:
        submission_sheet    = aef_sheet.AEFSubmissionSheet(workbook)
        submission_sheet.write_to_db(cursor, str_filepath)
    except Exception as e:
        # submission with same version already exists.  Mark as duplicate, move to duplicate directory,delete new submission from database.
        str_error	= "Duplicate submission.  Existing submission with the same Party ID (" + str_filepath[0:3] + "), Reported Year (" + str_filepath[4:8] + "), Version (" + str_filepath[9:10] + "." + str_filepath[11:12] + ") already exists"
        update_invalid_submission_version_worksheet_status(str_filepath, duplicate_dir, str_error)
    else:
        submission_key		    = submission_sheet.primary_key	# the submission primary key is used as the foreign key from other tables
        authorizations_sheet	= aef_sheet.AEFAuthorizationsSheet(workbook)
        authorizations_sheet.write_to_db(cursor, submission_key)
        actions_sheet			= aef_sheet.AEFActionsSheet(workbook)
        actions_sheet.write_to_db(cursor, submission_key)
        holdings_sheet			= aef_sheet.AEFHoldingsSheet(workbook)
        holdings_sheet.write_to_db(cursor, submission_key)
        auth_entities_sheet		= aef_sheet.AEFAuthEntitiesSheet(workbook)
        auth_entities_sheet.write_to_db(cursor, submission_key)	
    return


def replace_obsolete_submissions(cursor):
    """	Check new submissions in the database for consistency.
        New submissions are those with consistency_status IS NULL.
    """
    cursor.execute(f'SELECT party_id, reported_year, major_version, minor_version, path FROM Submissions WHERE consistency_status IS NULL')	# Get new submissions from db
    submission_rows	= cursor.fetchall()
    for submission_row in submission_rows:
        party_id, reported_year, major_version, minor_version, path = submission_row[0], submission_row[1], submission_row[2], submission_row[3], submission_row[4]
        cursor.execute(f'SELECT party_id, reported_year, major_version, minor_version, path FROM Submissions WHERE party_id="{party_id}" AND reported_year={reported_year} AND consistency_status IS NOT NULL')
        old_submission_rows	= cursor.fetchall()
        for old_submission_row in old_submission_rows:
            old_party_id, old_reported_year, old_major_version, old_minor_version, old_path = old_submission_row[0], old_submission_row[1], old_submission_row[2], old_submission_row[3], old_submission_row[4]
            if (old_major_version > major_version) or ((old_major_version == major_version) and (old_minor_version > minor_version)):
                # submission with newer version already exists.  Mark as obsolete, move to obsolete directory,delete new submission from database.
                str_error	= "Obsolete submission.  Existing submission with the same Party ID (" + party_id + "), Reported Year (" + reported_year + "), and later Version (" + old_major_version + "." + old_minor_version + ") already exists"
                update_invalid_submission_version_worksheet_status(path, obsolete_dir, str_error)
                delete_submission_from_db(cursor, party_id, reported_year, major_version, minor_version)
                break

			# if we reach here, submission is newer than old submission, so we move the old submission to "obsolete" directory and delete it from database
			# move old submission file to "obsolete" directory
            str_error	= "Obsoleted submission.  Newer submission with the same Party ID (" + party_id + "), Reported Year (" + reported_year + "), and later Version (" + major_version + "." + minor_version + ") submitted"
            update_invalid_submission_version_worksheet_status(old_path, obsolete_dir, str_error)
            delete_submission_from_db(cursor, old_party_id, old_reported_year, old_major_version, old_minor_version)
    cursor.connection.commit()
    return


def update_invalid_submission_version_worksheet_status(path, str_dir, str_error):
	"""	Handle the case where a submission has an invalid version number:
		either duplicate (the same as an existing submission) or obsolete (older than an existing submission).
	"""
	str_file	= os.path.basename(path)
	dst_path	= str_dir + str_file
	shutil.move(path, dst_path)							    # move source to destination
	workbook	= load_workbook(dst_path, data_only=True)	# load workbook openpyxl object
	worksheet	= workbook["Table 1 Submission"]
	worksheet.cell(9, 3, value=str_error + ". " + time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()) + " GMT")
	workbook.save(dst_path)
	return


def unload_inconsistent_submissions(cursor):
	"""	Delete the inconsistent submissions from the database.
	"""
	cursor.execute(f'SELECT party_id, reported_year, major_version, minor_version FROM Submissions WHERE consistency_status LIKE "%failed%"')	# Load inconsistent submissions from db
	submission_rows	= cursor.fetchall()
	for submission_row in submission_rows:
		party_id, reported_year, major_version, minor_version = submission_row[0], submission_row[1], submission_row[2], submission_row[3]
		delete_submission_from_db(cursor, party_id, reported_year, major_version, minor_version)
	cursor.connection.commit()
	return


def delete_submission_from_db(cursor, party_id, reported_year, major_version, minor_version):
	"""	Delete the specified submission from the database.
		Note that the operations are not committed here.
	"""
	cursor.execute(f'DELETE FROM Authorizations WHERE reporting_party_id="{party_id}" AND reported_year={reported_year} AND major_version={major_version} AND minor_version={minor_version}')
	cursor.execute(f'DELETE FROM Actions WHERE reporting_party_id="{party_id}" AND reported_year={reported_year} AND major_version={major_version} AND minor_version={minor_version}')
	cursor.execute(f'DELETE FROM Holdings WHERE reporting_party_id="{party_id}" AND reported_year={reported_year} AND major_version={major_version} AND minor_version={minor_version}')
	cursor.execute(f'DELETE FROM Authorized_Entities WHERE reporting_party_id="{party_id}" AND reported_year={reported_year} AND major_version={major_version} AND minor_version={minor_version}')
	cursor.execute(f'DELETE FROM Submissions WHERE party_id="{party_id}" AND reported_year={reported_year} AND major_version={major_version} AND minor_version={minor_version}')
	return


def check_new_submissions(cursor):
    """	Check new submissions in the database for consistency.
        New submissions are those with consistency_status IS NULL.
    """
    report	= AEFConsistencyReport()
    cursor.execute(f'SELECT * FROM Submissions WHERE consistency_status IS NULL')	# Load new submissions from db
    submission_rows	= cursor.fetchall()
    for submission_row in submission_rows:
        submission = aef_submission.AEFSubmission(cursor, submission_row)
        path       = submission.str_path
        str_file   = os.path.basename(path)
        report.reset()
        if (submission.is_consistent(cursor, report)):
            dst_path    = consistent_dir + str_file[0:13] + "consistent.xlsx"
            shutil.move(path, dst_path)							# move source to consistent directory
        else:
            dst_path    = inconsistent_dir + str_file[0:13] + "inconsistent.xlsx"
            shutil.move(path, dst_path)							# move source to inconsistent directory
        report.reset()
    return


def print_table(cursor, table_name):
	"""Print the contents of the specified table using the provided cursor.
	"""
	print(f'\n{table_name}:\n')
	cursor.execute(f'SELECT * FROM {table_name}')
	rows	= cursor.fetchall()
	for row in rows:    
		print(row)
	print("\n")
	return


def print_tables(cursor):
	"""Print the contents of all tables in the database using the provided cursor.
	"""
	tables	=	[
					"Submissions",
					"Authorizations",
					"Actions",
					"Holdings",
					"Authorized_Entities"
				]
	for table in tables:
		print_table(cursor, table)
	return


if __name__ == "__main__":
	main()
