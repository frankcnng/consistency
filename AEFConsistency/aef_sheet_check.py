import sqlite3
import re
from openpyxl import load_workbook


# aefsheets.py

from openpyxl.comments import Comment
import re
import sqlite3


class AEFSheetCheck:
    """Abstract superclass that checks the consistency of a worksheet."""

    def __init__(self, worksheet, labels):
        self.worksheet	= worksheet
        self.labels	    = labels
        return
    

class RowFieldsSheetCheck(AEFSheetCheck):
    """Abstract superclass for checks consistency of worksheets with fields organised in rows."""


class AEFAuthorizationsCheck(RowFieldsSheetCheck):
    """Concrete subclass to check AEF Authorizations worksheet."""

    def __init__(self, worksheet, field_names):
        self.template_sheet_name	= "Table 2 Authorizations"

        self.field_reg_exp_tuples	=	[
                                            ["Authorization ID", "[A-Za-z0-9 \-\/\.]+", "' can only contain alphanumeric, decimal point, space, hyphen, and slash characters."],
                                            ["Date of authorization", "dd/mm/yyyy", "' must be of the format 'dd/mm/yyyy'."],
                                            ["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
                                            ["Version of the authorization", "\d+", "' must be a number."],
                                            ["", "", ""],
                                            ["Authorized quantity", "\d+", "' must be a number."],
                                            ["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
                                            ["Applicable GWP value(s)", "", ""],
                                            ["Applicable non-GHG metric", "", ""],
                                            ["Sector(s)", "[A-Za-z0-9 ]+", "' can only contain alphanumeric, and space characters."],
                                            ["Activity type(s)", "[A-Za-z0-9 \+]+", "' can only contain alphanumeric, space, and '+' characters."],
                                            ["Purposes for authorization", "NDC|OIMP|IMP|OP|NDC and OIMP|NDC and IMP|NDC and OP", "' must be one of 'NDC', 'OIMP', 'IMP', 'OP', 'NDC and OIMP', 'NDC and IMP', or 'NDC and OP'."],
                                            ["Authorized Party(ies) ID", "blankable|NA|[A-Z]{3}( *, *[A-Z]{3})*", "' must a comma-separated list of ISO 3166 alpha-3 codes."],
                                            ["Authorized entity(ies) ID", "blankable|NA|[A-Za-z0-9 \-\.\(\)]+( *, *[A-Za-z0-9 \-\.\(\)]+)*", "' must be a comma-separated list of entity names."],
                                            ["OIMP authorized by the Party", "", ""],
                                            ["Authorized timeframe", "blankable|NA|^ *(?:Occurred|Use): from\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))\s+to\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))(?:(?:\.\s*Use:\s+from\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4}))\s+to\s+(?:(?:\d{4})|(?:\d{2}\/\d{4})|((?:0[1-9]|[12]\d|3[01])\/(?:0[1-9]|1[0-2])\/\d{4})))?) *$",\
                                                "' must be empty or 'Occurred: from <date> to <date>', 'Use: from <date> to <date>', or 'Occurred: from <date> to <date>. Use: from <date> to <date>'\n where <date> is 'yyyy', 'mm/yyyy', or 'dd/mm/yyyy'."],
                                            ["Authorization terms and conditions", "", ""],
                                            ["Authorization documentation", "", ""],
                                            ["First transfer definition for OIMP", "blankable|NA|Authorization|Issuance|Use or cancellation", "' must be empty or one of 'NA', 'Authorization', 'Issuance', 'Use or cancellation."],
                                            ["Additional explanatory information", "", ""]
                                        ]

        super().__init__(worksheet, field_names)


class AEFActionsCheck(RowFieldsSheetCheck):
    """Concrete subclass for AEF Actions worksheet."""

    def __init__(self, worksheet, field_names):
        self.template_sheet_name	= "Table 3 Actions"

        self.field_reg_exp_tuples	=	[
                                            ["Action date", "dd/mm/yyyy", "' must be of the format 'dd/mm/yyyy'."],
                                            ["Action type", "Acquisition|Transfer|Use|Cancellation|First transfer", "' must be one of 'Acquisition', 'Transfer', 'Use', 'Cancellation', 'First transfer'"],
                                            ["Action subtype", "", ""],
                                            ["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
                                            ["Authorization ID", "[A-Za-z0-9 \-\/\.]+", "' can only contain alphanumeric, decimal point, space, hyphen, and slash characters."],
                                            ["First transferring participating Party ID", "[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
                                            ["Party ITMO registry ID", "[A-Z]{3}\d{2}", "' must be a Party ID followed by two digits"],
                                            ["First ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
                                            ["Last ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
                                            ["", "", ""],
                                            ["Underlying unit registry ID", "", ""],
                                            ["First unit ID", "", ""],
                                            ["Last unit ID", "", ""],
                                            ["", "", ""],
                                            ["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
                                            ["Applicable GWP value(s)", "", ""],
                                            ["Applicable non-GHG metric", "", ""],
                                            ["Quantity (t CO2 eq)", "blankable|NA|\d+", "' must be a numerical value."],
                                            ["Quantity (in non-GHG metric)", "blankable|NA|\d*", "' must be a numerical value."],
                                            ["", "", ""],
                                            ["Mitigation type", "", ""],
                                            ["Vintage", "blankable|NA|\d{4}", "' must be a year."],
                                            ["", "", ""],
                                            ["Transferring participating Party ID", "blankable|NA|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
                                            ["Acquiring participating Party ID",  "blankable|NA|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
                                            ["", "", ""],
                                            ["Purpose for which the ITMO has been used towards or cancelled for OIMP", "", ""],
                                            ["Using/cancelling participating Party ID", "blankable|NA|[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
                                            ["Using/cancelling authorized entity ID", "", ""],
                                            ["Calendar year for which the ITMOs are used towards the Party's NDC", "blankable|NA|\d{4}", "' must be a year."],
                                            ["", "", ""],
                                            ["Result of the consistency checks", "", ""],
                                            ["Additional explanatory information", "", ""]
                                        ]

        super().__init__(worksheet, field_names)


class AEFHoldings(RowFieldsSheetCheck):
    """Concrete subclass for AEF Holdings worksheet."""

    def __init__(self, worksheet, field_names):
        self.template_sheet_name	= "Table 4 Holdings"

        self.field_reg_exp_tuples	=	[
                                            ["Cooperative approach ID", "CA\d{4}", "' must start with 'CA' followed by four digits."],
                                            ["Authorization ID", "[A-Za-z0-9 \-]+", "' can only contain alphanumeric, space, and hyphen characters."],
                                            ["First transferring participating Party ID", "[A-Z]{3}", "' must an ISO 3166 alpha-3 country code."],
                                            ["Party ITMO registry ID", "[A-Z]{3}\d{2}", "' must be a Party ID followed by two digits"],
                                            ["First ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
                                            ["Last ID", "CA\d{4}-[A-Z]{3}\d{2}-[A-Z]{3}-[1-9]\d{0,2}((\d*)|(,\d{3})*)-\d{4}", "' must be an ITMO unique identifier as per 6/CMA.4 annex I para.5."],
                                            ["", "", ""],
                                            ["Underlying unit registry ID", "", ""],
                                            ["First unit ID", "", ""],
                                            ["Last unit ID", "", ""],
                                            ["", "", ""],
                                            ["Metric", "GHG|non\-GHG", "' must 'GHG' or 'non-GHG'"],
                                            ["Applicable GWP value(s)", "", ""],
                                            ["Applicable non-GHG metric", "", ""],
                                            ["Quantity (t CO2 eq)", "blankable|NA|\d+", "' must be a numerical value."],
                                            ["Quantity (in non-GHG metric)", "blankable|NA|\d+", "' must be a numerical value."],
                                            ["", "", ""],
                                            ["Mitigation type", "", ""],
                                            ["Vintage", "\d{4}", "' must be a year."]
                                        ]

        super().__init__(worksheet, field_names)


class AEFAuthEntitiesCheck(RowFieldsSheetCheck):
    """Concrete subclass for AEF Authorized Entities worksheet."""

    def __init__(self, worksheet, field_names):
        self.template_sheet_name	= "Table 5 Auth. entities"

        self.field_reg_exp_tuples	=	[
                                            ["Date of the authorization", "blankable|dd/mm/yyyy", "' must be of the format 'dd/mm/yyyy'"],
                                            ["Name", "", ""],
                                            ["Country of incorporation", "blankable|[A-Z][A-Za-z \(\)\']+", "' is not a recognised Party Name."],	# Capitalised alphabet string contains spaces, brackets, apostrophes],
                                            ["Identification number", "", ""],
                                            ["Cooperative approach ID", "blankable|CA\d{4}", "' must start with 'CA' followed by four digits."],
                                            ["Conditions", ""],
                                            ["Change and revocation conditions", "", ""],
                                            ["Additional explanatory information", "", ""]
                                        ]

        super().__init__(worksheet, field_names)


class ColumnFieldsSheetCheck(AEFSheetCheck):
    """Abstract superclass for checking worksheets with fields organised in columns.
    """

    def get_field_dimensions(self):
        """Get the dimensions of the fields in the worksheet.
        """

        labels		    = self.labels
        nFields	        = len(labels)
        worksheet	    = self.worksheet
        str_heading     = labels[0].casefold()
        str_first_field = labels[1].casefold()
        str_last_field  = labels[nFields - 1].casefold()
        heading_row		= 0
        start_row	    = 0
        end_row		    = 0
        start_column	= 0
        end_column		= 0
        for row in worksheet.iter_cols(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
            for cell in row:
                if (cell.data_type == "s"):
                    if ((cell.value.casefold() == str_heading)):  # looking for the heading row
                        heading_row     = cell.row
                        start_column	= cell.column
                        end_column      = cell.column
                        continue
                    elif ((heading_row > 0) and (cell.value.casefold() == str_first_field)):    # looking for the first field label
                        start_row	= cell.row
                        continue
                    elif ((start_row > 0) and (cell.value.casefold() == str_last_field)):       # looking for the last field label
                        end_row	= cell.row
                        break
        return (start_row, start_column, end_row, end_column)
    

    def check_field_names(self, book_report):
        heading_row			= 0
        fields_start_row	= 0
        fields_end_row		= 0
        fields_column		= 0
        template_fields		= self.field_names
        n_template_fields	= len(template_fields)
        worksheet			= self.worksheet
        str_subheading		= self.template_sheet_name
        if (self.template_sheet_name == "Summary information"):
            str_subheading	+= ": " + template_fields[0]
        sheet_report		= AEFSheetReport(str_subheading, 3)
        book_report.add_sheet_report(sheet_report)
        for row in worksheet.iter_cols(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
            for cell in row:
                if ((cell.data_type == "s") and (cell.value.casefold() == template_fields[0].casefold())):
                    heading_row	= cell.row
                    fields_column	= cell.column
                    continue
                elif ((heading_row > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[1].casefold())):
                    fields_start_row	= cell.row
                    continue
                elif ((fields_start_row > 0) and (cell.data_type == "s") and (cell.value.casefold() == template_fields[n_template_fields - 1].casefold())):
                    fields_end_row	= cell.row
                    break
        if (self.check_fields_dimensions(heading_row, fields_start_row, fields_end_row, 0, worksheet, template_fields, sheet_report)) is False:
            return False

        dest_fields	= []
        for row in worksheet.iter_cols(min_col=fields_column, max_col=fields_column, min_row=fields_start_row, max_row=fields_end_row, values_only=True):
            dest_fields.append(row)

        for x_field in range (1, n_template_fields):
            if (template_fields[x_field] in dest_fields[0]) is False:
                sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "Could not find '" + template_fields[x_field] + "' field in worksheet.")
                return False
        sheet_report.add_cell_report(self.template_sheet_name, worksheet.cell(1,1), "All fields found.")
        return True



class AEFSubmissionCheck(ColumnFieldsSheetCheck):
    """Concrete subclass for AEF Submission worksheet.
    """
    def __init__(self, workbook):
        worksheet   = workbook["Table 1 Submission"]
        labels	=	[
                        "Table 1: Submission",
                        ["Party", "party_id"],
                        ["Version", "version"],
                        ["Reported year", "reported_year"],
                        ["Date of submission", "date_of_submission"],
                        ["Review status of the initial report", "consistency_status"],
                        ["First year of the NDC implementation period", "NDC_period_start_year"],
                        ["Last year of the NDC implementation period", "NDC_period_end_year"],
                        ["Reference to the Article 6 technical expert review report of the initial report", ""]
                    ]
        super().__init__(worksheet, labels)
        return

    def load_to_db(self, cursor):
        """Load the data from the worksheet into the database using the provided cursor.
        """
        
        worksheet    = self.worksheet
        start_row, start_column, end_row, end_column	= self.get_field_dimensions()

        party_id    = worksheet.cell(start_row, start_column).value
        version     = worksheet.cell(start_row + 1, start_column).value
        reported_year	= worksheet.cell(start_row + 2, start_column).value
        date_of_submission	= worksheet.cell(start_row + 3, start_column).value
        consistency_status	= worksheet.cell(start_row + 4, start_column).value
        NDC_period_start_year	= worksheet.cell(start_row + 5, start_column).value
        NDC_period_end_year	= worksheet.cell(start_row + 6, start_column).value

        cursor.execute("INSERT INTO Submissions (party_Id, version, reported_year, date_of_submission, consistency_status, NDC_period_start_year, NDC_period_end_year) VALUES (?, ?, ?, ?, ?, ?, ?)", (party_id, version, reported_year, date_of_submission, consistency_status, NDC_period_start_year, NDC_period_end_year))

        cursor.connection.commit()

        return
