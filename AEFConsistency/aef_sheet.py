# aef_sheet.py

from openpyxl import load_workbook

from openpyxl.comments import Comment
import re
import sqlite3


class AEFSheet:
    """Abstract superclass for an AEF worksheet within an AEF workbook.
    Did not make this a subclass of openpyxl.worksheet.Worksheet as they would require changing
    openpyxl code to modify how the sheets are created under openpyxl.workbook.Workbook.
    """

    def __init__(self, worksheet, labels):
        self.worksheet	= worksheet
        self.labels	    = labels
        return


class AEFRowFieldsSheet(AEFSheet):
    """Abstract superclass for AEF worksheets with fields organised in rows."""


    def get_field_dimensions(self):
        """Get the dimensions of the fields in the worksheet.
        """
        labels		    = self.labels
        nFields	        = len(labels)
        worksheet	    = self.worksheet
        str_heading     = labels[0][0].casefold()
        str_first_field = labels[1][0].casefold()
        str_last_field  = labels[nFields - 1][0].casefold()
        heading_row		= start_row     = end_row   = 0
        start_column    = end_column    = 0
        for row in worksheet.iter_rows(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
            is_blank_row    = True
            for cell in row:
                if (cell.value == None):
                    continue
                else:
                    is_blank_row    = False
                if (cell.data_type == "s"):
                    if ((cell.value.casefold() == str_heading)):                                # looking for the heading row
                        heading_row     = cell.row
                        start_column    = cell.column
                        continue
                    elif ((heading_row > 0) and (cell.value.casefold() == str_first_field)):    # looking for the first field label
                        start_row       = cell.row
                        start_column    = cell.column
                        continue
                    elif ((start_row > 0) and (cell.value.casefold() == str_last_field)):       # looking for the last field label
                        end_column	= cell.column
                        break
            if ((end_column > 0) and (is_blank_row) and (start_row > 0)):                                           # looking for the first fully blank row after the data
                end_row    = cell.row - 1
                break
        if (end_row == 0):
            end_row = worksheet.max_row
        return (start_row, start_column, end_row, end_column)


    def write_to_db(self, cursor, submission_key):
        """write the data from the worksheet to the database using the provided cursor.
        """
        worksheet   = self.worksheet
        labels      = self.labels
        start_row, start_column, end_row, end_column	= self.get_field_dimensions()
        start_row   += 1	# inital row of values is one past the labels row

        table_name  = labels[0][1]  # e.g., "Authorizations"

        col_names   = [label[1] for label in self.labels[1:]]       # e.g., ["auth_id", "date", ...]
        col_names   = [label for label in col_names if label]       # remove any empty labels
        fk_names    = ["reporting_party_Id", "reported_year", "major_version", "minor_version"]
        col_names.extend(fk_names)

        str_placeholders    = ", ".join("?" for _ in col_names)
        set_ITMO_id_labels  = {"First ID", "Last ID", "First unit ID", "Last unit ID"}
        for (x_row) in range(start_row, end_row + 1):
            x_label     = 0
            values      = []
            str_refs  = []
            for x_column in range(start_column, end_column + 1):
                x_label += 1
                str_label   = labels[x_label][0]
                if (str_label == ""):    # skip blank formatting columns
                    continue
                cell_value  = worksheet.cell(x_row, x_column).value
                if ((str_label in set_ITMO_id_labels) and (cell_value is not None)):
                    cell_value  = cell_value.replace(',', '')
                values.append(cell_value)

            str_col_names = ", ".join(f'{c}' for c in col_names)
            values.extend([submission_key["party_id"], submission_key["reported_year"], submission_key["major_version"], submission_key["minor_version"]])
            insert_sql = f'INSERT INTO {table_name} ({str_col_names}) VALUES ({str_placeholders})'
            cursor.execute(insert_sql, values)

        cursor.connection.commit()
        return


class AEFAuthorizationsSheet(AEFRowFieldsSheet):
    """Concrete subclass for AEF Authorizations worksheet.
    """

    def __init__(self, workbook):
        worksheet   = workbook["Table 2 Authorizations"]
        labels	    =	[
                            ["Table 2: Authorizations", "Authorizations"],
                            ["Authorization ID", "authorization_id"],
                            ["Date of authorization", "date_of_authorization"],
                            ["Cooperative approach ID", "cooperative_approach_id"],
                            ["Version of the authorization", "version_of_authorization"],
                            ["", ""],
                            ["Authorized quantity", "authorised_quantity"],
                            ["Metric", "metric"],
                            ["Applicable GWP value(s)", "applicable_gwp_values"],
                            ["Applicable non-GHG metric", "applicable_nonghg_metric"],
                            ["Sector(s)", "sectors"],
                            ["Activity type(s)", "activity_types"],
                            ["Purposes for authorization", "purposes_for_auth"],
                            ["Authorized Party(ies) ID", "authorized_parties"],
                            ["Authorized entity(ies) ID", "authorized_entities"],
                            ["OIMP authorized by the Party", "oimp_authorized"],
                            ["Authorized timeframe", "authorized_timeframe"],
                            ["Authorization terms and conditions", "auth_tcs"],
                            ["Authorization documentation", "auth_documentation"],
                            ["First transfer definition for OIMP", "first_transfer_defn4oimp"]
                        ]
        super().__init__(worksheet, labels)
        return


class AEFActionsSheet(AEFRowFieldsSheet):
    """Concrete subclass for AEF Actions worksheet.
    """

    def __init__(self, workbook):
        worksheet   = workbook["Table 3 Actions"]
        labels      =	[
                            ["Table 3: Actions", "Actions"],
                            ["Action date", "action_date"],
                            ["Action type", "action_type"],
                            ["Action subtype", "action_subtype"],
                            ["Cooperative approach ID", "cooperative_approach_id"],
                            ["Authorization ID", "authorization_id"],
                            ["First transferring participating Party ID", "first_transferring_party_id"],
                            ["Party ITMO registry ID", "party_itmo_registry_id"],
                            ["First ID", "first_id"],
                            ["Last ID", "last_id"],
                            ["", ""],
                            ["Underlying unit registry ID", "underlying_unit_registry_id"],
                            ["First unit ID", "first_unit_id"],
                            ["Last unit ID", "last_unit_id"],
                            ["", "", ""],
                            ["Metric", "metric"],
                            ["Applicable GWP value(s)", "applicable_gwp_values"],
                            ["Applicable non-GHG metric", "applicable_nonghg_metric"],
                            ["Quantity (t CO2 eq)", "quantity"],
                            ["Quantity (in non-GHG metric)", "quantity_nonghg_metric"],
                            ["", ""],
                            ["Mitigation type", "mitigation_type"],
                            ["Vintage", "vintage"],
                            ["", ""],
                            ["Transferring participating Party ID", "transferring_party_id"],
                            ["Acquiring participating Party ID",  "acquiring_party_id"],
                            ["", ""],
                            ["Purpose for which the ITMO has been used towards or cancelled for OIMP", "purpose"],
                            ["Using/cancelling participating Party ID", "use_cancelling_party_id"],
                            ["Using/cancelling authorized entity ID", "use_cancelling_entity_id"],
                            ["Calendar year for which the ITMOs are used towards the Party's NDC", "year_used_for_ndc"],
                            ["", ""],
                            ["Result of the consistency checks", "consistency_results"],
                            ["Additional explanatory information", "additional_info"]
                        ]
        super().__init__(worksheet, labels)
        return


class AEFHoldingsSheet(AEFRowFieldsSheet):
    """Concrete subclass for AEF Holdings worksheet.
    """

    def __init__(self, workbook):
        worksheet	= workbook["Table 4 Holdings"]
        labels      =	[
                            ["Table 4: Holdings", "Holdings"],
                            ["Cooperative approach ID", "cooperative_approach_id"],
                            ["Authorization ID", "authorization_id"],
                            ["First transferring participating Party ID", "first_transferring_party_id"],
                            ["Party ITMO registry ID", "party_itmo_registry_id"],
                            ["First ID", "first_id"],
                            ["Last ID", "last_id"],
                            ["", "", ""],
                            ["Underlying unit registry ID", "underlying_unit_registry_id"],
                            ["First unit ID", "first_unit_id"],
                            ["Last unit ID", "last_unit_id"],
                            ["", "", ""],
                            ["Metric", "metric"],
                            ["Applicable GWP value(s)", "applicable_gwp_values"],
                            ["Applicable non-GHG metric", "applicable_nonghg_metric"],
                            ["Quantity (t CO2 eq)", "quantity"],
                            ["Quantity (in non-GHG metric)", "quantity_nonghg_metric"],
                            ["", "", ""],
                            ["Mitigation type", "mitigation_type"],
                            ["Vintage", "vintage"]
                        ]
        super().__init__(worksheet, labels)
        return


class AEFAuthEntitiesSheet(AEFRowFieldsSheet):
    """Concrete subclass for AEF Authorized Entities worksheet.
    """

    def __init__(self, workbook):
        worksheet	= workbook["Table 5 Auth. entities"]
        labels  	=	[
                            ["Table 5: Authorized Entities", "Authorized_Entities"],
                            ["Date of the authorization", "date_of_authorization"],
                            ["Name", "name_of_entity"],
                            ["Country of incorporation", "country_of_incorporation"],
                            ["Identification number", "identification_number"],
                            ["Cooperative approach ID", "cooperative_approach_id"],
                            ["Conditions", "conditions"],
                            ["Change and revocation conditions", "change_revoc_conditions"],
                            ["Additional explanatory information", "additional_info"]
                        ]
        super().__init__(worksheet, labels)
        return


class AEFColumnFieldsSheet(AEFSheet):
    """Abstract superclass for AEF worksheets with fields organised in columns.
    """

    def get_field_dimensions(self):
        """Get the dimensions of the fields in the worksheet.
        Assumes there is only one column of data
        """
        labels		    = self.labels
        nFields	        = len(labels)
        worksheet	    = self.worksheet
        str_heading     = labels[0][0].casefold()
        str_first_field = labels[1][0].casefold()
        str_last_field  = labels[nFields - 1][0].casefold()
        heading_row		= start_row     = end_row   = 0
        start_column    = end_column    = 0
        for column in worksheet.iter_cols(min_row=worksheet.min_row, max_row=worksheet.max_row, min_col=worksheet.min_column, max_col=worksheet.max_column):
            for cell in column:
                if (cell.data_type == "s"):
                    if ((cell.value.casefold() == str_heading)):  # looking for the heading row
                        heading_row     = start_row     = end_row   = cell.row
                        start_column    = end_column                = cell.column
                        continue
                    elif ((heading_row > 0) and (cell.value.casefold() == str_first_field)):    # looking for the first field label
                        start_row       = end_row       = cell.row
                        start_column    = end_column    = cell.column
                        continue
                    elif ((start_row > 0) and (cell.value.casefold() == str_last_field)):       # looking for the last field label
                        end_row	    = cell.row
                        end_column	= cell.column
                        break
        return (start_row, start_column, end_row, end_column)
    

class AEFSubmissionSheet(AEFColumnFieldsSheet):
    """Concrete subclass for AEF Submission worksheet.
    """
    def __init__(self, workbook):
        worksheet   = workbook["Table 1 Submission"]
        labels	=	[
                        ["Table 1: Submission", "Submissions"],
                        ["Party", "party_id"],
                        ["Version", "major_version", "minor_version"],
                        ["Reported year", "reported_year"],
                        ["Date of submission", "date_of_submission"],
                        ["Review status of the initial report", "review_status"],
                        ["Result of the consistency check of this AEF submission", "consistency_status"],
                        ["First year of the NDC implementation period", "NDC_period_start_year"],
                        ["Last year of the NDC implementation period", "NDC_period_end_year"],
                        ["Reference to the Article 6 technical expert review report of the initial report", ""]
                    ]
        self.primary_key    = {
                                "party_id": "",
                                "reported_year": 0,
                                "major_version": 0,
                                "minor_version": 0
                            }
        super().__init__(worksheet, labels)
        return


    def write_to_db(self, cursor, str_path):
        """Write the data from the worksheet to the database using the provided cursor.
        """
        table_name  = self.labels[0][1]  # e.g., "Submissions"
        worksheet   = self.worksheet
        start_row, start_column, end_row, end_column	= self.get_field_dimensions()
        start_column += 1	# inital column of values is one past the labels column

        party_id                = worksheet.cell(start_row, start_column).value
        version                 = worksheet.cell(start_row + 1, start_column).value
        reported_year	        = worksheet.cell(start_row + 2, start_column).value
        date_of_submission	    = worksheet.cell(start_row + 3, start_column).value
        review_status           = worksheet.cell(start_row + 4, start_column).value
        consistency_status	    = worksheet.cell(start_row + 5, start_column).value
        ndc_period_start_year	= worksheet.cell(start_row + 6, start_column).value
        ndc_period_end_year	    = worksheet.cell(start_row + 7, start_column).value
    
        version_type    = type(version)
        if (version_type == int):
            major_version   = version
            minor_version   = 0
        elif (version_type == float):
            str_version = "{:.1f}".format(version)
            str_major_version, str_minor_version	= map(int, str_version.split('.'))
            major_version	= int(str_major_version)
            minor_version	= int(str_minor_version)
        else:
            str_version = str(version)
            if '.' in str_version:
                str_major_version, str_minor_version	= str_version.split('.')
                major_version	= int(str_major_version)
                minor_version	= int(str_minor_version)
            else:
                major_version	= int(str_version)
                minor_version	= 0

        if ((consistency_status == "{Information in this field is populated by the CARP}") or (consistency_status == "")) :
            consistency_status    = None

        cursor.execute("INSERT INTO Submissions (party_id, major_version, minor_version, reported_year, date_of_submission, review_status, consistency_status, ndc_period_start_year, ndc_period_end_year, path) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (party_id, major_version, minor_version, reported_year, date_of_submission, review_status, consistency_status, ndc_period_start_year, ndc_period_end_year, str_path))
        cursor.connection.commit()
        self.primary_key.update(  {
                                    "party_id": party_id,
                                    "reported_year": reported_year,
                                    "major_version": major_version,
                                    "minor_version": minor_version
                                }   )
        return
