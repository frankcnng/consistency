# AEFConsistencyReport.py

import time

# Load the openpyxl Excel library
from openpyxl import load_workbook
from openpyxl.styles import Font


class AEFConsistencyReport:
    """Class reponsible for producing the consistency check report for the AEF workbook."""

    def __init__(self):
        self.check_reports  = []
        self.str_success    = "Consistency checks passed"
        self.str_fail       = "Consistency checks failed"


    def add_check_report(self, check_report):
        self.check_reports.append(check_report)
        return True


    def reset(self):
          self.check_reports    = []


    def print(self, workbook, is_valid):
        str_title   = "Consistency check results"
        results_sheet	= workbook.create_sheet(str_title, index = 8)
        results_sheet.cell(1, 1, value=str_title)
        results_sheet.cell(1, 1).font	= Font(bold=True)

        x_row	= 2
        for check_report in self.check_reports:
            x_row	= check_report.print(workbook, results_sheet, x_row)
        x_row	+= 1

        if (is_valid):
            str_result  = self.str_success
        else:
            str_result  = self.str_fail
        cell        = results_sheet.cell(x_row, 2, value=str_result)
        cell.font   = Font(bold=True)

        if (is_valid):
            str_consistency = self.str_success
        else:
            str_consistency = self.str_fail
        worksheet   = workbook["Table 1 Submission"]
        worksheet.cell(9, 3, value=str_consistency + ": " + time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()) + " GMT")


class AEFCheckReport:
    """Class reponsible for producing the report for a consistency check."""

    def __init__(self, str_title):
        self.str_title      = str_title
        self.error_reports  = []


    def add_error_report(self, str_message):
        self.error_reports.append(str_message)


    def print(self, workbook, results_sheet, x_row):
        cell	        = results_sheet.cell(x_row, 2, value=self.str_title)
        cell.font       = Font(italic=True)
        x_row           += 1
        error_reports   = self.error_reports
        if (len(error_reports) == 0):
            cell	    = results_sheet.cell(x_row, 3, value="Check succeeded.")
            cell.font   = Font(italic=True)
            x_row       += 1
        else:
            list_reported   = []
            for error_report in error_reports:
                if (error_report in list_reported):
                    continue
                if (error_report[0] == '\t'):
                    results_sheet.cell(x_row, 5, value=error_report)
                else:
                    results_sheet.cell(x_row, 4, value=error_report)
                list_reported.append(error_report)
                x_row   += 1
            cell	    = results_sheet.cell(x_row, 3, value="Check failed.")
            cell.font   = Font(italic=True)
            x_row       += 1
        return (x_row)